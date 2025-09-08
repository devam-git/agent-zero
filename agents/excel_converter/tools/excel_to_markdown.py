import os
import openpyxl
from datetime import datetime
from python.helpers.files import get_abs_path
from python.helpers.tool import Tool, Response

class excel_to_markdown(Tool):

    async def execute(self, excel_path: str, output_filename: str = None):
        try:
            # Handle path
            if not os.path.isabs(excel_path):
                root_dir = get_abs_path("/root/")
                full_excel_path = os.path.join(root_dir, excel_path)
            else:
                full_excel_path = excel_path
            
            # Check file exists
            if not os.path.exists(full_excel_path):
                return Response(message=f"Excel file not found: {full_excel_path}", break_loop=False)
            
            # Load workbook with formula preservation
            try:
                # First try to load with formulas preserved
                workbook = openpyxl.load_workbook(full_excel_path, data_only=False, keep_vba=False)
            except Exception as e1:
                try:
                    # Fallback to data-only mode
                    workbook = openpyxl.load_workbook(full_excel_path, data_only=True)
                except Exception as e2:
                    return Response(message=f"Cannot load Excel file: {str(e1)} | Fallback error: {str(e2)}", break_loop=False)
            
            # Generate comprehensive markdown with formulas
            filename = os.path.basename(full_excel_path)
            markdown_content = self._generate_comprehensive_markdown(workbook, full_excel_path, filename)
            
            # Save or return
            if output_filename:
                save_dir = get_abs_path("/root/")
                if not output_filename.endswith('.md'):
                    output_filename += '.md'
                output_path = os.path.join(save_dir, output_filename)
                    
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(markdown_content)
                    
                return Response(message=f"Excel converted to Markdown: {output_path}", break_loop=False)
            else:
                return Response(message=f"Excel to Markdown conversion completed:\n\n{markdown_content}", break_loop=False)
                
        except Exception as e:
            return Response(message=f"Error converting Excel to Markdown: {str(e)}", break_loop=False)
    
    def _generate_comprehensive_markdown(self, workbook, file_path, filename):
        """Comprehensive markdown generation with formula extraction"""
        markdown = f"# Excel File: {filename}\n\n"
        
        # Metadata
        markdown += "## Metadata\n"
        markdown += f"- **File Path**: {file_path}\n"
        markdown += f"- **Processing Date**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        markdown += f"- **Total Sheets**: {len(workbook.worksheets)}\n\n"
        
        # Process each sheet with formula extraction
        for sheet in workbook.worksheets:
            markdown += f"## Sheet: {sheet.title}\n\n"
            
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            if max_row == 1 and max_col == 1 and sheet.cell(1, 1).value is None:
                markdown += "*(Empty sheet)*\n\n"
                continue
            
            # Sheet summary
            formulas = []
            cell_details = []
            non_empty_cells = 0
            
            # Collect all cell data, formulas, and details
            for row in sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    if cell.value is not None:
                        non_empty_cells += 1
                        cell_coord = f"{cell.coordinate}"
                        
                        # Check if cell contains a formula
                        if str(cell.value).startswith('='):
                            formula_text = str(cell.value)
                            try:
                                # Try to get calculated value
                                calculated_value = cell.displayed_value if hasattr(cell, 'displayed_value') else "Not calculated"
                            except:
                                calculated_value = "Error in calculation"
                            
                            formulas.append({
                                'coord': cell_coord,
                                'formula': formula_text,
                                'result': calculated_value
                            })
                            cell_details.append(f"- **{cell_coord}**: Formula → `{formula_text}` (Result: {calculated_value})")
                        else:
                            # Regular value
                            cell_type = "Unknown"
                            if isinstance(cell.value, str):
                                cell_type = "Text"
                            elif isinstance(cell.value, (int, float)):
                                cell_type = "Number"
                            elif hasattr(cell.value, 'date'):
                                cell_type = "Date"
                            
                            cell_details.append(f"- **{cell_coord}**: {cell_type} → {str(cell.value)[:50]}")
            
            # Summary section
            markdown += "### Summary\n"
            markdown += f"- **Dimensions**: {max_row} x {max_col}\n"
            markdown += f"- **Data Range**: A1:{sheet.cell(max_row, max_col).coordinate}\n"
            markdown += f"- **Formula Count**: {len(formulas)}\n"
            markdown += f"- **Non-empty Cells**: {non_empty_cells}\n\n"
            
            # Data table (first 10 rows x 8 cols)
            markdown += "### Data Table\n\n"
            display_rows = min(10, max_row)
            display_cols = min(8, max_col)
            
            if display_rows > 0 and display_cols > 0:
                # Header row
                headers = []
                for col in range(1, display_cols + 1):
                    cell = sheet.cell(1, col)
                    headers.append(str(cell.value or f"Col{col}").replace("|", "\\|")[:20])
                
                markdown += "| " + " | ".join(headers) + " |\n"
                markdown += "|" + "---|" * len(headers) + "\n"
                
                # Data rows
                for row_num in range(2, display_rows + 1):
                    row_data = []
                    for col in range(1, display_cols + 1):
                        cell = sheet.cell(row_num, col)
                        if str(cell.value or "").startswith('='):
                            # Show formula with result
                            try:
                                result = cell.displayed_value if hasattr(cell, 'displayed_value') else "Not calculated"
                                cell_display = f"[Formula] Result: {result}"
                            except:
                                cell_display = f"[Formula: {str(cell.value)[:20]}]"
                        else:
                            cell_display = str(cell.value or "")
                        
                        row_data.append(cell_display.replace("|", "\\|")[:30])
                    
                    markdown += "| " + " | ".join(row_data) + " |\n"
                
                if max_row > 10:
                    markdown += f"\n*(... and {max_row - 10} more rows)*\n"
            
            markdown += "\n"
            
            # Formulas detected section
            if formulas:
                markdown += "### Formulas Detected\n"
                for formula in formulas[:20]:  # Show first 20 formulas
                    markdown += f"- **{formula['coord']}**: `{formula['formula']}` → Result: {formula['result']}\n"
                
                if len(formulas) > 20:
                    markdown += f"*(... and {len(formulas) - 20} more formulas)*\n"
                markdown += "\n"
            
            # Cell details section (first 50 cells)
            if cell_details:
                markdown += "### Cell Details\n"
                for detail in cell_details[:50]:
                    markdown += detail + "\n"
                
                if len(cell_details) > 50:
                    markdown += f"*(... and {len(cell_details) - 50} more cells)*\n"
                markdown += "\n"
            
            markdown += "---\n\n"
        
        return markdown